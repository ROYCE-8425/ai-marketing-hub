"""
Keyword Rank Tracker — Phase 10 (Upgraded to SQLAlchemy)

Tracks keyword ranking positions over time using:
1. Google Search Console API (free, already configured)
2. DataForSEO API (optional, when configured)

Features:
- PostgreSQL history with trend charts
- Tag/group keywords by campaign
- CSV import/export
- Ranking drop alerts
"""

import os
import io
import csv
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List
import httpx

from dotenv import load_dotenv
from core.database import SessionLocal
from core.models import ManagedSite, TrackedKeyword, RankingHistory


def _get_site_id(db, site_url: str) -> int:
    site = db.query(ManagedSite).filter(ManagedSite.url == site_url).first()
    if not site:
        raise ValueError(f"Site {site_url} không tồn tại trong hệ thống")
    return site.id


def add_keyword(keyword: str, site_url: str, tag: str = "") -> Dict[str, Any]:
    """Add a keyword to track."""
    db = SessionLocal()
    try:
        site_id = _get_site_id(db, site_url)
        kw_text = keyword.strip().lower()
        
        existing = db.query(TrackedKeyword).filter(
            TrackedKeyword.keyword == kw_text,
            TrackedKeyword.site_id == site_id
        ).first()
        
        if not existing:
            kw = TrackedKeyword(keyword=kw_text, site_id=site_id, tag=tag.strip())
            db.add(kw)
            db.commit()
            
        return {"status": "ok", "keyword": kw_text, "tag": tag.strip()}
    except ValueError as e:
        return {"status": "error", "message": str(e)}
    finally:
        db.close()


def update_keyword_tag(keyword: str, site_url: str, tag: str) -> Dict[str, Any]:
    """Update tag for a tracked keyword."""
    db = SessionLocal()
    try:
        site_id = _get_site_id(db, site_url)
        kw = db.query(TrackedKeyword).filter(
            TrackedKeyword.keyword == keyword.strip().lower(),
            TrackedKeyword.site_id == site_id
        ).first()
        
        if kw:
            kw.tag = tag.strip()
            db.commit()
        return {"status": "ok", "keyword": keyword.strip().lower(), "tag": tag.strip()}
    except ValueError as e:
        return {"status": "error", "message": str(e)}
    finally:
        db.close()


def remove_keyword(keyword: str, site_url: str) -> Dict[str, Any]:
    """Remove a tracked keyword."""
    db = SessionLocal()
    try:
        site_id = _get_site_id(db, site_url)
        kw_text = keyword.strip().lower()
        kw = db.query(TrackedKeyword).filter(
            TrackedKeyword.keyword == kw_text,
            TrackedKeyword.site_id == site_id
        ).first()
        
        if kw:
            db.delete(kw)
            db.commit()
        return {"status": "ok", "removed": kw_text}
    except ValueError as e:
        return {"status": "error", "message": str(e)}
    finally:
        db.close()


def get_tracked_keywords(site_url: str, tag_filter: str = "") -> List[Dict[str, Any]]:
    """Get all tracked keywords with latest ranking."""
    db = SessionLocal()
    try:
        site = db.query(ManagedSite).filter(ManagedSite.url == site_url).first()
        if not site:
            return []

        # We will fetch keywords and then manually find their latest 2 rankings 
        # (SQL window functions are trickier in basic SQLAlchemy ORM, raw SQL was used before)
        
        query = db.query(TrackedKeyword).filter(TrackedKeyword.site_id == site.id)
        if tag_filter:
            query = query.filter(TrackedKeyword.tag == tag_filter)
            
        keywords = query.all()
        results = []
        
        for kw in keywords:
            # Get latest 2 rankings
            history = db.query(RankingHistory).filter(RankingHistory.tracked_keyword_id == kw.id)\
                        .order_by(RankingHistory.checked_at.desc()).limit(2).all()
            
            latest = history[0] if len(history) > 0 else None
            prev = history[1] if len(history) > 1 else None
            
            pos = latest.position if latest else None
            prev_pos = prev.position if prev else None
            change = None
            if pos is not None and prev_pos is not None:
                change = round(prev_pos - pos, 1)

            results.append({
                "keyword": kw.keyword,
                "tag": kw.tag or "",
                "position": pos,
                "previous_position": prev_pos,
                "change": change,
                "clicks": latest.clicks if latest else 0,
                "impressions": latest.impressions if latest else 0,
                "ctr": latest.ctr if latest else 0,
                "source": latest.source if latest else "pending",
                "last_checked": latest.checked_at.isoformat() if latest and latest.checked_at else None,
                "created_at": kw.created_at.isoformat() if kw.created_at else None,
            })
            
        # Sort by position ASC NULLS LAST
        results.sort(key=lambda x: (x["position"] is None, x["position"]))
        return results
    finally:
        db.close()


def get_all_tags(site_url: str) -> List[str]:
    """Get all unique tags."""
    db = SessionLocal()
    try:
        site = db.query(ManagedSite).filter(ManagedSite.url == site_url).first()
        if not site:
            return []
            
        tags = db.query(TrackedKeyword.tag).filter(
            TrackedKeyword.site_id == site.id,
            TrackedKeyword.tag != ""
        ).distinct().all()
        return [t[0] for t in tags]
    finally:
        db.close()


def import_keywords_csv(csv_text: str, site_url: str) -> Dict[str, Any]:
    """Import keywords from CSV text."""
    db = SessionLocal()
    try:
        site_id = _get_site_id(db, site_url)
        reader = csv.reader(io.StringIO(csv_text.strip()))
        added = 0
        skipped = 0
        for row in reader:
            if not row or not row[0].strip():
                continue
            kw_text = row[0].strip().lower()
            tag = row[1].strip() if len(row) > 1 else ""
            if kw_text in ("keyword", "từ khóa", "tu khoa"):
                continue
            
            existing = db.query(TrackedKeyword).filter(
                TrackedKeyword.keyword == kw_text,
                TrackedKeyword.site_id == site_id
            ).first()
            
            if not existing:
                kw = TrackedKeyword(keyword=kw_text, site_id=site_id, tag=tag)
                db.add(kw)
                added += 1
            else:
                skipped += 1
        db.commit()
        return {"status": "ok", "added": added, "skipped": skipped}
    except ValueError as e:
        return {"status": "error", "message": str(e)}
    finally:
        db.close()


def export_keywords_csv(site_url: str) -> str:
    """Export tracked keywords with latest rankings as CSV."""
    keywords = get_tracked_keywords(site_url)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Keyword", "Tag", "Vị trí", "Thay đổi", "Clicks", "Hiển thị", "CTR", "Nguồn", "Cập nhật"])
    for kw in keywords:
        writer.writerow([
            kw["keyword"], kw.get("tag", ""),
            kw["position"] or "—", kw["change"] or "—",
            kw["clicks"], kw["impressions"],
            f"{kw['ctr']}%" if kw["ctr"] else "—",
            kw["source"], kw["last_checked"] or "—",
        ])
    return output.getvalue()


def export_keywords_excel(site_url: str) -> bytes:
    """Export tracked keywords with latest rankings as styled XLSX bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl>=3.1.0")

    keywords = get_tracked_keywords(site_url)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rank Tracker"

    headers = ["Keyword", "Tag", "Vị trí", "Thay đổi", "Clicks", "Hiển thị", "CTR (%)", "Nguồn", "Cập nhật"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="8B5CF6")
    thin = Side(style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for row_idx, kw in enumerate(keywords, start=2):
        pos = kw.get("position")
        change = kw.get("change")
        values = [
            kw.get("keyword", ""),
            kw.get("tag", ""),
            pos if pos else "—",
            change if change is not None else "—",
            kw.get("clicks", 0),
            kw.get("impressions", 0),
            kw.get("ctr", 0),
            kw.get("source", ""),
            kw.get("last_checked") or "—",
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = border
            c.alignment = center
            if col == 3 and isinstance(v, (int, float)):
                if v <= 3: c.font = Font(color="16A34A", bold=True)
                elif v <= 10: c.font = Font(color="2563EB", bold=True)
                elif v <= 20: c.font = Font(color="D97706", bold=True)
                else: c.font = Font(color="DC2626", bold=True)
            if col == 4 and isinstance(v, (int, float)):
                if v > 0: c.font = Font(color="16A34A")
                elif v < 0: c.font = Font(color="DC2626")

    widths = [32, 16, 10, 12, 10, 12, 10, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def check_ranking_alerts(site_url: str, threshold: int = 5) -> List[Dict[str, Any]]:
    """Check for keywords that dropped more than threshold positions."""
    keywords = get_tracked_keywords(site_url)
    alerts = []
    for kw in keywords:
        if kw["change"] is not None and kw["change"] < -threshold:
            alerts.append({
                "keyword": kw["keyword"],
                "current_position": kw["position"],
                "previous_position": kw["previous_position"],
                "drop": abs(kw["change"]),
                "severity": "critical" if kw["change"] < -10 else "warning",
            })
    return alerts


def get_keyword_history(keyword: str, site_url: str, days: int = 30) -> List[Dict[str, Any]]:
    """Get ranking history for a keyword."""
    db = SessionLocal()
    try:
        site_id = _get_site_id(db, site_url)
        kw = db.query(TrackedKeyword).filter(
            TrackedKeyword.keyword == keyword.strip().lower(),
            TrackedKeyword.site_id == site_id
        ).first()
        
        if not kw:
            return []
            
        since = datetime.now(timezone.utc) - timedelta(days=days)
        history = db.query(RankingHistory).filter(
            RankingHistory.tracked_keyword_id == kw.id,
            RankingHistory.checked_at >= since
        ).order_by(RankingHistory.checked_at.asc()).all()

        return [
            {
                "date": r.checked_at.isoformat()[:10] if r.checked_at else "",
                "position": r.position,
                "clicks": r.clicks,
                "impressions": r.impressions,
                "ctr": r.ctr,
                "source": r.source,
            }
            for r in history
        ]
    except ValueError:
        return []
    finally:
        db.close()


def save_ranking(keyword: str, site_url: str, position: float,
                 clicks: int = 0, impressions: int = 0, ctr: float = 0,
                 source: str = "gsc") -> None:
    """Save a ranking data point."""
    db = SessionLocal()
    try:
        site_id = _get_site_id(db, site_url)
        kw = db.query(TrackedKeyword).filter(
            TrackedKeyword.keyword == keyword.strip().lower(),
            TrackedKeyword.site_id == site_id
        ).first()
        
        if not kw:
            return

        rh = RankingHistory(
            tracked_keyword_id=kw.id,
            position=position,
            clicks=clicks,
            impressions=impressions,
            ctr=ctr,
            source=source
        )
        db.add(rh)
        db.commit()
    except ValueError:
        pass
    finally:
        db.close()


async def sync_rankings_from_gsc(site_url: str) -> Dict[str, Any]:
    """Fetch current rankings from Google Search Console and save to DB."""
    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env"), override=True)
    client_id = os.getenv("GOOGLE_SEARCH_CONSOLE_CLIENT_ID", "")
    client_secret = os.getenv("GOOGLE_SEARCH_CONSOLE_CLIENT_SECRET", "")
    refresh_token = os.getenv("GOOGLE_SEARCH_CONSOLE_REFRESH_TOKEN", "")
    gsc_site = os.getenv("GSC_SITE_URL", site_url)

    if not all([client_id, client_secret, refresh_token]):
        return {"error": "GSC chưa cấu hình OAuth2", "synced": 0}

    async with httpx.AsyncClient(timeout=15) as client:
        token_resp = await client.post(
            "https://oauth2.googleapis.com/token",
            data={
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
            },
        )
        if token_resp.status_code != 200:
            return {"error": "Không thể refresh token GSC", "synced": 0}

        access_token = token_resp.json().get("access_token")

        db = SessionLocal()
        try:
            site_id = _get_site_id(db, site_url)
            tracked = db.query(TrackedKeyword).filter(TrackedKeyword.site_id == site_id).all()
            if not tracked:
                return {"error": "Chưa có keyword nào được theo dõi", "synced": 0}

            keywords = [k.keyword for k in tracked]
        except ValueError as e:
            return {"error": str(e), "synced": 0}
        finally:
            db.close()

        synced_count = 0
        for kw in keywords:
            try:
                resp = await client.post(
                    f"https://www.googleapis.com/webmasters/v3/sites/{gsc_site}/searchAnalytics/query",
                    headers={"Authorization": f"Bearer {access_token}"},
                    json={
                        "startDate": (datetime.utcnow() - timedelta(days=3)).strftime("%Y-%m-%d"),
                        "endDate": datetime.utcnow().strftime("%Y-%m-%d"),
                        "dimensions": ["query"],
                        "dimensionFilterGroups": [{
                            "filters": [{
                                "dimension": "query",
                                "expression": kw,
                                "operator": "equals",
                            }]
                        }],
                        "rowLimit": 1,
                    },
                )
                if resp.status_code == 200:
                    data = resp.json()
                    rows = data.get("rows", [])
                    if rows:
                        row = rows[0]
                        save_ranking(
                            keyword=kw,
                            site_url=site_url,
                            position=round(row.get("position", 0), 1),
                            clicks=row.get("clicks", 0),
                            impressions=row.get("impressions", 0),
                            ctr=round(row.get("ctr", 0) * 100, 2),
                            source="gsc",
                        )
                        synced_count += 1
                    else:
                        save_ranking(kw, site_url, 0, source="gsc_not_found")
                        synced_count += 1
            except Exception:
                continue

        return {"status": "ok", "synced": synced_count, "total_keywords": len(keywords)}
